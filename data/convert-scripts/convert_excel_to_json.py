#!/usr/bin/env python3
"""
Konvertiert die SRZ_Regelfragen_Erfassung.xlsx in das JSON-Format
für die Schiedsrichter-Trainingsapp.

Usage:
    python convert_excel_to_json.py <excel_file> [--output <json_file>] [--append-to <existing_json>]

Beispiele:
    python convert_excel_to_json.py ../../SRZ_Regelfragen_Erfassung.xlsx
    python convert_excel_to_json.py ../../SRZ_Regelfragen_Erfassung.xlsx --output questions-new.json
    python convert_excel_to_json.py ../../SRZ_Regelfragen_Erfassung.xlsx --append-to questions-all.json
"""

import json
import re
import sys
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl nicht installiert. Bitte 'pip install openpyxl' ausführen.")
    sys.exit(1)


# === CRITERIA EXTRACTION ===
CRITERIA_KEYWORDS = [
    "Strafstoß", "Direkter Freistoß", "Indirekter Freistoß",
    "Eckstoß", "Einwurf", "Abstoß", "Anstoß",
    "Feldverweis", "Verwarnung", "Gelb/Rot", "Gelb-Rot",
    "Spielabbruch", "Spielunterbrechung", "Spielende",
    "Wiederholung", "Schiedsrichterball",
    "Tor", "Kein Tor",
    "Ja", "Nein",
    "Vorteil",
]

PARTIAL_PRIORITY = [
    "Feldverweis", "Gelb/Rot", "Verwarnung",
    "Strafstoß", "Direkter Freistoß", "Indirekter Freistoß",
    "Spielabbruch", "Tor", "Kein Tor",
]

TAG_PATTERNS = {
    "Abseits": [r"[Aa]bseits"],
    "Handspiel": [r"[Hh]andspiel", r"mit der Hand", r"mit dem Arm"],
    "Foulspiel": [r"[Ff]oulspiel", r"[Hh]altevergehen", r"tritt.*gegen", r"[Bb]einstellen"],
    "Strafstoß": [r"[Ss]trafstoß", r"[Ee]lfmeter"],
    "Torwart": [r"[Tt]orwart", r"[Tt]orhüter", r"[Tt]orsteher"],
    "Persönliche Strafe": [r"[Vv]erwarnung", r"[Ff]eldverweis", r"Gelb.?Rot", r"Rote Karte", r"Gelbe Karte"],
    "Spielfortsetzung": [r"[Ff]reistoß", r"[Ss]trafstoß", r"[Ee]inwurf", r"[Ee]ckstoß", r"[Aa]bstoß", r"[Aa]nstoß", r"Schiedsrichterball"],
    "Notbremse/DOGSO": [r"[Nn]otbremse", r"DOGSO", r"klare Torchance", r"offensichtliche Torchance"],
    "Unsportliches Verhalten": [r"[Uu]nsportlich", r"[Ss]imulation", r"[Zz]eitverzögerung"],
    "Tätlichkeit": [r"[Tt]ätlichkeit", r"[Ss]chlagen", r"[Ss]pucken", r"[Bb]eißen"],
    "VAR": [r"\bVAR\b", r"[Vv]ideo.?[Aa]ssist"],
    "Auswechslung": [r"[Aa]uswechsl", r"[Ee]inwechsl"],
    "Ausrüstung": [r"[Aa]usrüstung", r"[Ss]chmuck", r"[Ss]chienbeinschoner"],
    "Elfmeterschießen": [r"[Ee]lfmeterschieß"],
    "Verlängerung": [r"[Vv]erlängerung", r"[Nn]achspielzeit"],
    "Vorteil": [r"[Vv]orteil"],
    "Schiedsrichter-Entscheidung": [r"[Ss]chiedsrichter.*entscheid", r"[Ss]chiedsrichter.*Irrtum"],
    "Spielfeld": [r"[Ss]pielfeld", r"[Tt]orlinie", r"[Ss]eitenlinie", r"[Mm]ittelkreis"],
    "Allgemein": [],
}


def extract_criteria_full(answer: str) -> list[str]:
    """Extrahiert criteriaFull aus der Antwort."""
    found = []
    answer_lower = answer.lower()
    first_sentence = answer.split(".")[0] + "." if "." in answer else answer

    # Check first sentence and full answer for keywords
    for kw in CRITERIA_KEYWORDS:
        if kw.lower() in first_sentence.lower() or kw.lower() in answer_lower:
            # Special handling: "Ja"/"Nein" only from first word
            if kw in ("Ja", "Nein"):
                first_word = answer.strip().split()[0].rstrip(".,!") if answer.strip() else ""
                if first_word == kw:
                    found.append(kw)
            elif kw == "Gelb-Rot":
                if "Gelb/Rot" not in found:
                    found.append("Gelb/Rot")
            elif kw == "Tor" and kw.lower() in answer_lower:
                # Avoid false positive from "Torwart", "Torhüter", etc.
                if re.search(r'\b[Tt]or\b(?!wart|hüter|linie|raum|schuss)', answer):
                    if kw not in found:
                        found.append(kw)
            elif kw == "Kein Tor":
                if re.search(r'[Kk]ein\s+Tor\b', answer):
                    if kw not in found:
                        found.append(kw)
            else:
                if kw not in found:
                    found.append(kw)
    return found


def extract_criteria_partial(criteria_full: list[str]) -> list[str]:
    """Leitet criteriaPartial aus criteriaFull ab (wichtigste Kriterien)."""
    partial = []
    for kw in PARTIAL_PRIORITY:
        if kw in criteria_full:
            partial.append(kw)
            break  # Nur das wichtigste
    if not partial and criteria_full:
        # Fallback: erstes Nicht-Ja/Nein Kriterium
        for c in criteria_full:
            if c not in ("Ja", "Nein"):
                partial.append(c)
                break
    return partial


def extract_tags(situation: str, answer: str) -> list[str]:
    """Extrahiert Tags aus Situation und Antwort."""
    combined = situation + " " + answer
    tags = []
    for tag, patterns in TAG_PATTERNS.items():
        if tag == "Allgemein":
            continue
        for pattern in patterns:
            if re.search(pattern, combined):
                tags.append(tag)
                break
    if not tags:
        tags.append("Allgemein")
    return sorted(tags)


def source_to_date(source_type: str, ausgabe: str) -> str:
    """Konvertiert Quellentyp + Ausgabe zu sourceDate (YYYY-MM-01)."""
    parts = ausgabe.strip().split("/")
    if len(parts) != 2:
        raise ValueError(f"Ungültiges Ausgabenformat: '{ausgabe}' (erwartet MM/YYYY)")
    month_str, year_str = parts
    try:
        month = int(month_str)
        year = int(year_str)
    except ValueError:
        raise ValueError(f"Ungültiges Ausgabenformat: '{ausgabe}' (Monat/Jahr nicht numerisch)")
    if not (1 <= month <= 12):
        raise ValueError(f"Ungültiger Monat: {month} in '{ausgabe}'")
    if not (2013 <= year <= 2030):
        raise ValueError(f"Ungültiges Jahr: {year} in '{ausgabe}'")
    return f"{year}-{month:02d}-01"


def validate_row(row_num: int, quellentyp: str, ausgabe: str, situation: str, answer: str) -> list[str]:
    """Validiert eine Zeile und gibt Fehlerliste zurück."""
    errors = []
    if quellentyp not in ("SR-Zeitung", "SR-Newsletter"):
        errors.append(f"Zeile {row_num}: Ungültiger Quellentyp '{quellentyp}' (erlaubt: SR-Zeitung, SR-Newsletter)")
    if not re.match(r'^\d{2}/\d{4}$', str(ausgabe).strip()):
        errors.append(f"Zeile {row_num}: Ungültiges Ausgabenformat '{ausgabe}' (erwartet: MM/YYYY)")
    if not situation or len(str(situation).strip()) < 10:
        errors.append(f"Zeile {row_num}: Situation fehlt oder zu kurz")
    if not answer or len(str(answer).strip()) < 5:
        errors.append(f"Zeile {row_num}: Antwort fehlt oder zu kurz")
    return errors


def convert_excel_to_json(excel_path: str, start_index: int = 1) -> tuple[list[dict], list[str]]:
    """
    Liest die Excel-Datei und konvertiert in JSON-Format.
    Returns: (questions_list, errors_list)
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Regelfragen']

    # Detect header row and column mapping
    headers = {}
    for col in range(1, 20):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[val.strip()] = col

    # Map to expected columns
    col_map = {}
    for expected, alternatives in {
        'quellentyp': ['Quellentyp', 'Quelle'],
        'ausgabe': ['Ausgabe'],
        'situation': ['Situation'],
        'antwort': ['Antwort'],
        'regelreferenz': ['Regelreferenz'],
    }.items():
        for alt in alternatives:
            if alt in headers:
                col_map[expected] = headers[alt]
                break

    # Check if old format (single "Quelle" column with "SR-Zeitung 03/2025")
    old_format = 'quellentyp' not in col_map and 'Quelle' in headers and 'ausgabe' not in col_map

    if old_format:
        print("INFO: Altes Tabellenformat erkannt (einzelne Quelle-Spalte)")
        col_quelle = headers['Quelle']
    else:
        required = ['quellentyp', 'ausgabe', 'situation', 'antwort']
        missing = [r for r in required if r not in col_map]
        if missing:
            return [], [f"Fehlende Spalten: {', '.join(missing)}"]

    questions = []
    errors = []
    idx = start_index

    for row in range(2, ws.max_row + 1):
        # Read values
        if old_format:
            quelle_raw = ws.cell(row=row, column=col_quelle).value
            situation = ws.cell(row=row, column=headers.get('Situation', 3)).value
            answer = ws.cell(row=row, column=headers.get('Antwort', 4)).value
            regelref = ws.cell(row=row, column=headers.get('Regelreferenz', 5)).value
        else:
            quellentyp = ws.cell(row=row, column=col_map.get('quellentyp')).value
            ausgabe = ws.cell(row=row, column=col_map.get('ausgabe')).value
            situation = ws.cell(row=row, column=col_map['situation']).value
            answer = ws.cell(row=row, column=col_map['antwort']).value
            regelref = ws.cell(row=row, column=col_map.get('regelreferenz')).value if 'regelreferenz' in col_map else None

        # Skip empty rows (check Situation as primary indicator)
        if not situation or str(situation).strip() == '':
            continue

        situation = str(situation).strip()
        answer = str(answer).strip() if answer else ''

        # Parse source
        if old_format:
            if not quelle_raw:
                errors.append(f"Zeile {row}: Quelle fehlt")
                continue
            quelle_str = str(quelle_raw).strip()
            match = re.match(r'^(SR-Zeitung|SR-Newsletter)\s+(\d{2}/\d{4})$', quelle_str)
            if not match:
                errors.append(f"Zeile {row}: Ungültiges Quellenformat '{quelle_str}' (erwartet: 'SR-Zeitung MM/YYYY')")
                continue
            quellentyp = match.group(1)
            ausgabe = match.group(2)
        else:
            quellentyp = str(quellentyp).strip() if quellentyp else ''
            ausgabe = str(ausgabe).strip() if ausgabe else ''

        # Validate
        row_errors = validate_row(row, quellentyp, ausgabe, situation, answer)
        if row_errors:
            errors.extend(row_errors)
            continue

        # Build source string
        source = f"{quellentyp} {ausgabe}"

        # Parse sourceDate
        try:
            source_date = source_to_date(quellentyp, ausgabe)
        except ValueError as e:
            errors.append(f"Zeile {row}: {e}")
            continue

        # Extract criteria and tags
        criteria_full = extract_criteria_full(answer)
        criteria_partial = extract_criteria_partial(criteria_full)
        tags = extract_tags(situation, answer)

        # Build explanation (same as answer, with rule reference appended)
        explanation = answer
        if regelref and str(regelref).strip():
            regelref_str = str(regelref).strip()
            if regelref_str not in explanation:
                explanation += f" (Vgl. {regelref_str})"

        question = {
            "index": idx,
            "situation": situation,
            "correctAnswer": answer,
            "source": source,
            "criteriaFull": criteria_full,
            "criteriaPartial": criteria_partial,
            "sourceDate": source_date,
            "ruleReference": str(regelref).strip() if regelref else "",
            "tags": tags,
            "explanation": explanation,
        }
        questions.append(question)
        idx += 1

    return questions, errors


def main():
    parser = argparse.ArgumentParser(description='Konvertiert SRZ Excel-Erfassung zu JSON')
    parser.add_argument('excel_file', help='Pfad zur Excel-Datei')
    parser.add_argument('--output', '-o', help='Ausgabe-JSON-Datei (Standard: questions-manual.json)')
    parser.add_argument('--append-to', help='An bestehende JSON-Datei anhängen')
    args = parser.parse_args()

    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"ERROR: Datei nicht gefunden: {excel_path}")
        sys.exit(1)

    # Determine start index
    start_index = 1
    existing_questions = []
    if args.append_to:
        append_path = Path(args.append_to)
        if append_path.exists():
            with open(append_path, 'r', encoding='utf-8') as f:
                existing_questions = json.load(f)
            start_index = max(q['index'] for q in existing_questions) + 1
            print(f"Bestehende Datei: {len(existing_questions)} Fragen (nächster Index: {start_index})")

    # Convert
    questions, errors = convert_excel_to_json(str(excel_path), start_index)

    # Report errors
    if errors:
        print(f"\n⚠️  {len(errors)} Fehler gefunden:")
        for e in errors:
            print(f"  - {e}")
        print()

    if not questions:
        print("Keine gültigen Fragen gefunden.")
        sys.exit(1 if errors else 0)

    # Determine output path
    if args.append_to:
        output_path = Path(args.append_to)
        all_questions = existing_questions + questions
    else:
        output_path = Path(args.output) if args.output else excel_path.parent / 'questions-manual.json'
        all_questions = questions

    # Write JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_questions, f, ensure_ascii=False, indent=2)

    print(f"✅ {len(questions)} Fragen konvertiert → {output_path}")
    if args.append_to:
        print(f"   Gesamt: {len(all_questions)} Fragen")

    # Summary
    sources = {}
    for q in questions:
        sources[q['source']] = sources.get(q['source'], 0) + 1
    print(f"\nQuellen:")
    for src, count in sorted(sources.items()):
        print(f"  {src}: {count} Fragen")


if __name__ == '__main__':
    main()
